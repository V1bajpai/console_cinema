from openpyxl import Workbook, load_workbook
from datetime import datetime, date



wb = Workbook()
sheet = wb.active


''' check whether user is customer or Admin '''
def isuser(value):
    if value == 1:
        pwd = input('Enter password: ')
        if pwd == '1234':
            adminfun()

    if value == 2:
        customerview()

    else:
        print('Wrong Choice')
        return

def adminfun():
    sheet.cell(row=1, column=1).value = 'Id'
    sheet.cell(row=1, column=2).value = 'Movie'
    sheet.cell(row=1, column=3).value = 'Slot'
    sheet.cell(row=1, column=4).value = 'Booked Seats'
    sheet.cell(row=1, column=5).value = 'Available Seats'
    sheet.cell(row=1, column=6).value = 'Booked'
    sheet.cell(row=1, column=7).value = 'cancel'
    sheet.cell(row=1, column=8).value = 'Ticket number'
    for i in range(2,6):
        for j in range(1,9):
            sheet.cell(row=i, column=j).value = input()
        print('Do you want to break press-0 to break')
        if int(input()) == 0:
            break
        else:
            continue


    wb.save(filename='test2.xlsx')
    wb.close()

def bookseats(seats, slot):
    now = datetime.now()
    current_time = now.strftime("%H:%M:%S")
    today = date.today()
    d1 = today.strftime("%d/%m/%Y")
    updated_seats = 100 - seats
    if updated_seats<=0:
        print('Ticket is not available')
        return
    ticket=[]
    for i in range(1,seats+1):
        ticket.append(str(i)+'_'+current_time+'_'+d1)

    wb = load_workbook("C:\\console_cinema\\test2.xlsx")
    ws = wb['Sheet']
    ws.cell(row=slot, column=6).value = 'Booked'
    ws.cell(row=slot, column=4).value = seats
    ws.cell(row=slot, column=5).value = updated_seats
    ws.cell(row=slot, column=8).value = str(ticket)
    wb.save("C:\\console_cinema\\test2.xlsx")
    wb.close()
    print('Successfully booked')


def cancelseats(slot):
    wb = load_workbook("C:\\console_cinema\\test2.xlsx")
    ws = wb['Sheet']
    cancelseats=int(input('How many seats you want to cancel-'))
    bookedseats = ws.cell(row=slot, column=4).value - cancelseats
    updated_seats = ws.cell(row=slot, column=5).value + cancelseats
    updated_ticket = ws.cell(row=slot, column=8).value
    updated_ticket = str(updated_ticket)[1:]
    updated_ticket = updated_ticket[:-1]
    updated_ticket = updated_ticket.replace("'", '')
    updated_ticket = updated_ticket.split(', ', )
    if updated_seats != 100:
        ws.cell(row=slot, column=6).value = 'Booked'
        ws.cell(row=slot, column=7).value = 'Cancelled'
    else:
        ws.cell(row=slot, column=6).value = ''
        ws.cell(row=slot, column=7).value = 'Cancelled'
    ws.cell(row=slot, column=4).value = bookedseats
    ws.cell(row=slot, column=5).value = updated_seats
    ws.cell(row=slot, column=8).value = str(updated_ticket[cancelseats:])
    wb.save("C:\\console_cinema\\test2.xlsx")
    wb.close()
    print('Successfully Cancelled')


def customerview():
    wb1 = load_workbook(filename="test2.xlsx")
    ws = wb1.active
    for i in range(1, 6):
        for j in range(1, 9):
            print(ws.cell(row=i, column=j).value, end='   |   ')
        print('\n')

    print("What do you want to do- ")
    print("Press-1: For booking seats")
    print("Press-2: For Cancelling seats")
    customerchoice=int(input())
    if customerchoice == 1:
        print("Book Seats here- ")
        seats=int(input("Enter the number of seats: "))
        slot=int(input('Enter the slot:'))
        bookseats(seats, slot+1)

    if customerchoice == 2:
        slot = int(input('Enter the slot:'))
        cancelseats(slot+1)


print('Choose your profile: ')
print('Press 1: Admin')
print('Press 2: Customer')
isuser(int(input()))



